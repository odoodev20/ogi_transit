import base64
import io
from odoo import models, fields, api, _
from odoo.exceptions import ValidationError
from markupsafe import Markup

try:
    import openpyxl
except ImportError:
    openpyxl = None

class OgiPlImportWizard(models.TransientModel):
    _name = 'ogi.pl.import.wizard'
    _description = 'Packing List Import Wizard'

    container_id = fields.Many2one('ogi.transit.container', string='Container', required=True)
    import_file = fields.Binary(string='Excel File', required=True)
    file_name = fields.Char(string='File Name')
    
    # Warning tracking fields
    has_bgda_warning = fields.Boolean(default=False)
    ignore_bgda_warning = fields.Boolean(default=False)

    @api.onchange('import_file')
    def _onchange_import_file(self):
        self.has_bgda_warning = False
        self.ignore_bgda_warning = False

    def _clean_float(self, val):
        """Helper to handle Excel numbers formatted with spaces (e.g. 38 900) or commas"""
        if not val:
            return 0.0
        if isinstance(val, (int, float)):
            return float(val)
        clean_str = str(val).replace(' ', '').replace(',', '')
        try:
            return float(clean_str)
        except ValueError:
            return 0.0

    def action_force_import(self):
        self.ignore_bgda_warning = True
        return self.action_import_excel()

    def action_import_excel(self):
        if not openpyxl:
            raise ValidationError("The 'openpyxl' Python library is not installed on the server.")
        
        # Decode the uploaded file
        file_content = base64.b64decode(self.import_file)
        data = io.BytesIO(file_content)
        workbook = openpyxl.load_workbook(data, data_only=True)
        sheet = workbook.active
        
        # Check for BGDA column header (Row 8, Column 8)
        bgda_header = str(sheet.cell(row=8, column=8).value or '').strip().upper()
        if 'BGDA' not in bgda_header and not self.ignore_bgda_warning:
            self.has_bgda_warning = True
            return {
                'type': 'ir.actions.act_window',
                'res_model': 'ogi.pl.import.wizard',
                'res_id': self.id,
                'view_mode': 'form',
                'target': 'new',
            }

        try:
            # 1. Read Global Financial Headers (Vertical layout in Column B/2)
            total_freight_usd = self._clean_float(sheet.cell(row=2, column=2).value)
            sales_price_gnf = self._clean_float(sheet.cell(row=4, column=2).value)
            ff_cost_gnf = self._clean_float(sheet.cell(row=5, column=2).value)
        except Exception as e:
            raise ValidationError(f"Failed to read the financial totals at the top of the Excel sheet. {str(e)}")

        # 2. Parse Lines and Apply Deduplication Logic
        lines_to_create = []
        warnings = []
        Partner = self.env['res.partner']
        
        # Track total CBM for validation
        total_imported_cbm = 0.0
        
        # NEW: Dictionary to track phone numbers within the file to prevent duplicates
        seen_phones = {}

        # Start reading from Row 9 downwards based on the new template
        for row_idx in range(9, sheet.max_row + 1):
            customer_name = sheet.cell(row=row_idx, column=1).value
            phone = sheet.cell(row=row_idx, column=2).value

            # Skip empty rows
            if not customer_name and not phone:
                continue

            # Strict Rule: Phone is mandatory
            if not phone:
                raise ValidationError(f"Import BLOCKED: Missing phone number on row {row_idx}. Phone number is the mandatory deduplication key.")

            # Clean the phone number of any accidental spaces
            phone_str = str(phone).replace(" ", "").strip()
            customer_name_str = str(customer_name).strip()

            # NEW: Intra-file duplication check
            if phone_str in seen_phones:
                raise ValidationError(
                    f"Import BLOCKED: Duplicate phone number '{phone_str}' detected on row {row_idx} "
                    f"(previously seen on row {seen_phones[phone_str]}). "
                    f"Please consolidate all goods for this customer into a single row before importing."
                )
            
            # Log the phone number and its row for future duplicate checks
            seen_phones[phone_str] = row_idx

            # Deduplication Logic against the existing Database
            partner = Partner.search([('phone_1', '=', phone_str)], limit=1)

            if partner:
                if partner.name.lower() != customer_name_str.lower():
                    # This fulfills TC-US4.3-02: It links to the existing partner but logs a warning!
                    warnings.append(f"Row {row_idx}: Linked to existing customer {partner.name} (Phone: {phone_str}) despite name mismatch '{customer_name_str}'.")
            else:
                partner = Partner.create({
                    'name': customer_name_str,
                    'phone_1': phone_str,
                })

            # Calculate and track CBM
            cbm_value = self._clean_float(sheet.cell(row=row_idx, column=6).value)
            total_imported_cbm += cbm_value

            # Read line data using clean_float for numeric safety
            lines_to_create.append({
                'container_id': self.container_id.id,
                'partner_id': partner.id,
                'mark': str(sheet.cell(row=row_idx, column=3).value or ''),
                'goods_description': str(sheet.cell(row=row_idx, column=4).value or ''),
                'qty': self._clean_float(sheet.cell(row=row_idx, column=5).value),
                'cbm_line': cbm_value,
                'ins_fee': self._clean_float(sheet.cell(row=row_idx, column=7).value),
                'bgda': self._clean_float(sheet.cell(row=row_idx, column=8).value),
            })
            
        # Dubai & China Origin CBM Validation
        if self.container_id.origin == 'dubai':
            if round(total_imported_cbm, 2) != 43.00:
                raise ValidationError(f"Validation Error: Dubai origin containers must have exactly 43.0 Total CBM/Line. The imported file has a total of {round(total_imported_cbm, 2)} CBM.")
        elif self.container_id.origin == 'china':
            if round(total_imported_cbm, 2) != 68.00:
                raise ValidationError(f"Validation Error: China origin containers must have exactly 68.0 Total CBM/Line. The imported file has a total of {round(total_imported_cbm, 2)} CBM.")

        # 3. Security Check: Prevent import if payments exist
        if self.container_id.pl_line_ids.filtered(lambda l: (l.usd_invoice_id and l.usd_invoice_id.state in ['partial', 'paid']) or
                                                          (l.gnf_invoice_id and l.gnf_invoice_id.state in ['partial', 'paid'])):
            raise ValidationError("Import Blocked: You cannot re-import a Packing List for a container that already has processed payments.")

        # 4. Wipe old lines and insert new data
        self.container_id.pl_line_ids.unlink()
        self.env['ogi.transit.pl.line'].create(lines_to_create)

        # 5. Update Container Master Financials
        self.container_id.write({
            'total_freight_usd': total_freight_usd,
            'total_customs_gnf': sales_price_gnf,
            'total_freight_forwarder_gnf': ff_cost_gnf,
            'state': 'created' # Triggers the UI to unlock the calculation buttons
        })

        # 6. Log traceabilities and warnings to the chatter
        log_msg = f"<strong>Packing List Imported</strong><br/>{len(lines_to_create)} lines processed."
        if warnings:
            log_msg += "<br/><br/><strong>Warnings:</strong><ul>"
            for w in warnings:
                log_msg += f"<li>{w}</li>"
            log_msg += "</ul>"

        self.container_id.message_post(body=Markup(log_msg))

        return {'type': 'ir.actions.act_window_close'}